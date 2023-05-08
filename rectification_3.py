from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import (QCoreApplication, QObject, QRunnable, QThread,
                          QThreadPool, pyqtSignal )
from PyQt5.QtWidgets import *#QGraphicsObject,QLineEdit,QWidget,QHBoxLayout
from PyQt5.QtGui import QIcon
import socket
import os
import sys

from openpyxl import Workbook
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as xlimg

import qrcode
from PIL import Image
import json

import time
from datetime import date
from datetime import datetime,timedelta

import win32api, win32con
from win32com import client


import threading
from threading import Thread,Event



class Ui_FARICATION(QMainWindow):
    
    send = pyqtSignal(str)
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon('logo1.png'))

        
        
    def setupUi(self, FARICATION):
        FARICATION.setObjectName("FARICATION")
        FARICATION.resize(1139, 574)
        self.centralwidget = QtWidgets.QWidget(FARICATION)
        self.centralwidget.setObjectName("centralwidget")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(21, 11, 1091, 526))
        self.widget.setObjectName("widget")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.widget)
        self.gridLayout_3.setContentsMargins(0, 0, 0, 0)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.groupBox = QtWidgets.QGroupBox(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(14)
        self.groupBox.setFont(font)
        self.groupBox.setFlat(False)
        self.groupBox.setEnabled(False)
        self.groupBox.setObjectName("groupBox")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.groupBox)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.listWidget = QtWidgets.QListWidget(self.groupBox)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.listWidget.setFont(font)
        self.listWidget.setObjectName("listWidget")
        self.gridLayout.addWidget(self.listWidget, 0, 0, 1, 3)
        self.label_5 = QtWidgets.QLabel(self.groupBox)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.gridLayout.addWidget(self.label_5, 1, 0, 1, 1)
        
        self.lineEdit_4 = QtWidgets.QLineEdit(self.groupBox)
        #self.lineEdit_4.returnPressed.conncet(self.keyPressEvent)
        
        self.lineEdit_4.returnPressed.connect(self.recieve)
        
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.lineEdit_4.setFont(font)
        self.lineEdit_4.setObjectName("lineEdit_4")
        # set default message
        self.lineEdit_4.setText("bismi Allah")
        
       
        
        self.gridLayout.addWidget(self.lineEdit_4, 1, 1, 1, 1)
        
        self.Send = QtWidgets.QPushButton(self.groupBox)#,clicked = lambda:self.keyPressEvent())

        ########
        self.Send.clicked.connect(self.recieve)
        ########
        
        font = QtGui.QFont()
        font.setPointSize(12)
        self.Send.setFont(font)
        self.Send.setObjectName("Send")
        self.gridLayout.addWidget(self.Send, 1, 2, 1, 1)
        self.verticalLayout_3.addLayout(self.gridLayout)
        self.gridLayout_3.addWidget(self.groupBox, 0, 0, 2, 1)
        self.groupBox_4 = QtWidgets.QGroupBox(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(14)
        self.groupBox_4.setFont(font)
        self.groupBox_4.setObjectName("groupBox_4")
        self.verticalLayout_4 = QtWidgets.QVBoxLayout(self.groupBox_4)
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.formLayout = QtWidgets.QFormLayout()
        self.formLayout.setObjectName("formLayout")
        self.label_4 = QtWidgets.QLabel(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.label_4)
        self.lineEdit_1 = QtWidgets.QLineEdit(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.lineEdit_1.setFont(font)
        self.lineEdit_1.setObjectName("lineEdit_1")

        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.lineEdit_1)
        self.label_15 = QtWidgets.QLabel(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_15.setFont(font)
        self.label_15.setObjectName("label_15")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.label_15)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.groupBox_4)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.lineEdit_2)

        
        self.Send_2 = QtWidgets.QPushButton(self.groupBox_4,clicked = lambda:self.operator_demo())
        
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(12)
        self.Send_2.setFont(font)
        self.Send_2.setObjectName("Send_2")

        
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.FieldRole, self.Send_2)
        self.verticalLayout_4.addLayout(self.formLayout)
        self.gridLayout_3.addWidget(self.groupBox_4, 0, 1, 1, 1)
        self.groupBox_3 = QtWidgets.QGroupBox(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(14)
        self.groupBox_3.setFont(font)
        
        self.groupBox_3.setObjectName("groupBox_3")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.groupBox_3)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.label_9 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.verticalLayout_2.addWidget(self.label_9)
        self.label_6 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.verticalLayout_2.addWidget(self.label_6)
        self.label_12 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_12.setFont(font)
        self.label_12.setObjectName("label_12")
        self.verticalLayout_2.addWidget(self.label_12)
        self.label_11 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_11.setFont(font)
        self.label_11.setObjectName("label_11")
        self.verticalLayout_2.addWidget(self.label_11)

        self.label_10 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_10.setFont(font)
        self.label_10.setObjectName("label_10")
        self.verticalLayout_2.addWidget(self.label_10)


        self.label_16 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_16.setFont(font)
        self.label_16.setObjectName("label_16")
        self.verticalLayout_2.addWidget(self.label_16)



        
        self.label_8 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.verticalLayout_2.addWidget(self.label_8)
        self.label_13 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_13.setFont(font)
        self.label_13.setObjectName("label_13")
        self.verticalLayout_2.addWidget(self.label_13)
        self.label_14 = QtWidgets.QLabel(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_14.setFont(font)
        self.label_14.setObjectName("label_14")
        self.verticalLayout_2.addWidget(self.label_14)
        self.horizontalLayout.addLayout(self.verticalLayout_2)
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.nemero_fab = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.nemero_fab.setFont(font)
        self.nemero_fab.setObjectName("nemero_fab")
        self.verticalLayout.addWidget(self.nemero_fab)
        self.fournisseur = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.fournisseur.setFont(font)
        self.fournisseur.setObjectName("fournisseur")
        self.verticalLayout.addWidget(self.fournisseur)
        self.bobine = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.bobine.setFont(font)
        self.bobine.setObjectName("bobine")
        self.verticalLayout.addWidget(self.bobine)
        self.commande = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.commande.setFont(font)
        self.commande.setObjectName("commande")
        self.verticalLayout.addWidget(self.commande)
        
        self.diametre = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.diametre.setFont(font)
        self.diametre.setObjectName("diametre")
        self.verticalLayout.addWidget(self.diametre)

        self.date = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.date.setFont(font)
        self.date.setObjectName("date")
        self.verticalLayout.addWidget(self.date)

        
        self.installation = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.installation.setFont(font)
        self.installation.setObjectName("installation")
        self.verticalLayout.addWidget(self.installation)
        self.coulee = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.coulee.setFont(font)
        self.coulee.setObjectName("coulee")
        self.verticalLayout.addWidget(self.coulee)
        self.fournisseur_2 = QtWidgets.QLineEdit(self.groupBox_3)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(16)
        self.fournisseur_2.setFont(font)
        self.fournisseur_2.setObjectName("fournisseur_2")
        self.verticalLayout.addWidget(self.fournisseur_2)
        self.horizontalLayout.addLayout(self.verticalLayout)
        self.gridLayout_3.addWidget(self.groupBox_3, 1, 1, 1, 1)
        
        self.label_7 = QtWidgets.QLabel(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(12)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.gridLayout_3.addWidget(self.label_7, 2, 1, 1, 1)

        self.label_17 = QtWidgets.QLabel(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(12)
        self.label_17.setFont(font)
        self.label_17.setObjectName("label_17")
        self.gridLayout_3.addWidget(self.label_7, 2, 1, 1, 1)
        
        self.groupBox_2 = QtWidgets.QGroupBox(self.widget)
        font = QtGui.QFont()
        font.setFamily("Algerian")
        font.setPointSize(14)
        self.groupBox_2.setFont(font)
        self.groupBox_2.setFlat(False)
        #self.groupBox_2.setCheckable(True)
        self.groupBox_2.setChecked(True)
        self.groupBox_2.setObjectName("groupBox_2")
        self.verticalLayout_5 = QtWidgets.QVBoxLayout(self.groupBox_2)
        self.verticalLayout_5.setObjectName("verticalLayout_5")
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.comboBox = QtWidgets.QComboBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(12)
        self.comboBox.setFont(font)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.gridLayout_2.addWidget(self.comboBox, 0, 1, 1, 1)
        self.label = QtWidgets.QLabel(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.gridLayout_2.addWidget(self.label, 0, 0, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.checkBox_4 = QtWidgets.QCheckBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.checkBox_4.setFont(font)
        self.checkBox_4.setObjectName("checkBox_4")
        self.horizontalLayout_2.addWidget(self.checkBox_4)
        self.checkBox_5 = QtWidgets.QCheckBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.checkBox_5.setFont(font)
        self.checkBox_5.setObjectName("checkBox_5")
        self.horizontalLayout_2.addWidget(self.checkBox_5)
        self.checkBox_6 = QtWidgets.QCheckBox(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.checkBox_6.setFont(font)
        self.checkBox_6.setObjectName("checkBox_6")
        self.horizontalLayout_2.addWidget(self.checkBox_6)
        self.gridLayout_2.addLayout(self.horizontalLayout_2, 5, 0, 1, 2)
        self.label_3 = QtWidgets.QLabel(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.gridLayout_2.addWidget(self.label_3, 3, 0, 1, 1)
        self.textEdit = QtWidgets.QTextEdit(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.textEdit.setFont(font)
        self.textEdit.setObjectName("textEdit")
        
        self.gridLayout_2.addWidget(self.textEdit, 2, 0, 1, 2)
        self.label_2 = QtWidgets.QLabel(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.gridLayout_2.addWidget(self.label_2, 1, 0, 1, 1)
        
        self.Conform = QtWidgets.QPushButton(self.groupBox_2,clicked = lambda:self.confirmation()) 

        self.Conform.setEnabled(True)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.Conform.setFont(font)
        #self.Conform.setCheckable(False)
        self.Conform.setObjectName("Conform")
        
        self.gridLayout_2.addWidget(self.Conform, 6, 0, 1, 2)
        self.textEdit_2 = QtWidgets.QTextEdit(self.groupBox_2)
        font = QtGui.QFont()
        font.setFamily("Calibri")
        font.setPointSize(14)
        self.textEdit_2.setFont(font)
        self.textEdit_2.setObjectName("textEdit_2")
        self.gridLayout_2.addWidget(self.textEdit_2, 4, 0, 1, 2)
        self.verticalLayout_5.addLayout(self.gridLayout_2)
        self.gridLayout_3.addWidget(self.groupBox_2, 0, 2, 2, 1)
        FARICATION.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(FARICATION)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1139, 21))
        self.menubar.setObjectName("menubar")
        FARICATION.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(FARICATION)
        self.statusbar.setObjectName("statusbar")
        FARICATION.setStatusBar(self.statusbar)

        
        self.groupBox.setEnabled(False)
        self.groupBox_2.setEnabled(False)
        self.groupBox_3.setEnabled(False)
        
        self.lineEdit_1.setFocus()

        
        self.lineEdit_1.returnPressed.connect(self.set_focus)
        self.lineEdit_2.returnPressed.connect(self.operator_demo)
        

        self.retranslateUi(FARICATION)
        QtCore.QMetaObject.connectSlotsByName(FARICATION)

    def retranslateUi(self, FARICATION):
        _translate = QtCore.QCoreApplication.translate
        FARICATION.setWindowTitle(_translate("FARICATION", "DEMANDE DE RECTIFICATION"))
        self.groupBox.setTitle(_translate("FARICATION", "Méssgerie"))
        self.label_5.setText(_translate("FARICATION", "Message:"))
        self.Send.setText(_translate("FARICATION", "Send"))
        self.groupBox_4.setTitle(_translate("FARICATION", "Operateur Fabrication:"))
        self.label_4.setText(_translate("FARICATION", "Nom et prénom:"))
        self.label_15.setText(_translate("FARICATION", "Code Operateur:"))
        self.Send_2.setText(_translate("FARICATION", "Enter"))
        self.groupBox_3.setTitle(_translate("FARICATION", "TUBE INFO"))
        self.label_9.setText(_translate("FARICATION", "N° FABRICATION:"))
        self.label_6.setText(_translate("FARICATION", "FOURNISSEUR: "))
        self.label_12.setText(_translate("FARICATION", "N° BOBINE:"))
        
        self.label_11.setText(_translate("FARICATION", "N° DE COMMANDE:"))

        self.label_16.setText(_translate("FARICATION", "DATE:"))
        
        self.label_10.setText(_translate("FARICATION", "DIAMETRE:"))
        self.label_8.setText(_translate("FARICATION", "INSTALLATION:"))
        self.label_13.setText(_translate("FARICATION", "N° COULEE:"))
        self.label_14.setText(_translate("FARICATION", "EPAISSEUR:"))
        self.label_7.setText(_translate("FARICATION", "   Developped by bouzid yassine L2-RT: 2022"))
        self.groupBox_2.setTitle(_translate("FARICATION", "Confirmation"))
        self.comboBox.setItemText(0, _translate("FARICATION", "Main d\'oeuver"))
        self.comboBox.setItemText(1, _translate("FARICATION", "Matière"))
        self.comboBox.setItemText(2, _translate("FARICATION", "Milieu"))
        self.comboBox.setItemText(3, _translate("FARICATION", "Méthode"))
        self.comboBox.setItemText(4, _translate("FARICATION", "Machine"))
        self.label.setText(_translate("FARICATION", "Anlyse des causes:"))
        self.checkBox_4.setText(_translate("FARICATION", "Mise en oeuver"))
        self.checkBox_5.setText(_translate("FARICATION", "Efficacité"))
        self.checkBox_6.setText(_translate("FARICATION", "Cloture"))
        self.label_3.setText(_translate("FARICATION", "Action corrective:"))
        self.label_2.setText(_translate("FARICATION", "Action curative:"))
        self.Conform.setText(_translate("FARICATION", "CONFIRMER"))

    def set_focus(self):
         self.lineEdit_2.setFocus()

    def operator_demo(self):
        if self.lineEdit_1.text()!="" and self.lineEdit_2.text()=="":       
            self.groupBox.setEnabled(True)
            self.groupBox_2.setEnabled(True)
            self.groupBox_3.setEnabled(True)
            self.lineEdit_4.setFocus()
        else:
            self.groupBox.setEnabled(False)
            self.groupBox_2.setEnabled(False)
            self.groupBox_3.setEnabled(False)
            self.lineEdit_1.setFocus()

            
        print("demo!!")
        
  
    #recieve()
    def keyPressEvent(self, event):
        #if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Tab:
        print('enter event detected ')
        if event.key() == Qt.Key_Return:
            self.recieve()
        

    def recieve(self):
        chat.messge_recieved.connect(self.updateReceivedMessage)
        self.send.emit(str(self.lineEdit_4.text()))
        self.listWidget.insertItem(-1,"FAB: "+self.lineEdit_4.text())
        self.lineEdit_4.setText("")
    def updateReceivedMessage(self,txt):
        if txt!="":
            pass
            #it = QtGui.QStandardItem("RX1: "+txt)
            #self.listWidget.addItem("RX1: "+txt)
            #self.model.appendRow(it)
            #chat.quit()

            
    def checked(self):
         if self.checkBox_4.isChecked()==True:
             self.MISE_EN_OEUVER = "Oui"
         else :
             self.MISE_EN_OEUVER = "Non"
             
         if self.checkBox_5.isChecked()==True:
             self.EFFICACITE = "Oui"
         else :
             self.EFFICACITE = "Non"
             
         if self.checkBox_6.isChecked()==True:
             self.CLOTURE = "Oui"
         else :
             self.CLOTURE = "Non"
             
    def rq_genrator(self,string):
            logo = Image.open('logo1.png')
            basewidth = 150
             
            # adjust image size
            wpercent = (basewidth/float(logo.size[0]))
            hsize = int((float(logo.size[1])*float(wpercent)))
            logo = logo.resize((basewidth, hsize), Image.ANTIALIAS)
            QRcode = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H,
                                   box_size=20,
                                   border = 0)
            QRcode.add_data(string)
            QRcode.make()
            QRcolor = 'black'
            QRimg = QRcode.make_image(fill_color=QRcolor, back_color="white").convert('RGB')
            pos = ((QRimg.size[0] - logo.size[0]) // 2,
                   (QRimg.size[1] - logo.size[1]) // 2)
            QRimg.paste(logo, pos)
            new_width  = 170
            new_height = 135
            QRimg = QRimg.resize((new_width, new_height), Image.Resampling.LANCZOS)
            QRimg.save('img2.png')
            return QRimg


    def converting_excel_to_pdf(self,excel_path,pdf_path):
        
        app = client.Dispatch("Excel.Application")
        #app.Interactive=False
        app.Visible= False
        pdf_wb = app.Workbooks.Open(excel_path)
        pdf_wb.ActiveSheet.ExportAsFixedFormat(0,pdf_path)
        pdf_wb.Close()
            
    
    def confirmation(self):
        self.checked()
        self.rq_genrator("Operateur: "+self.lineEdit_1.text()+"  date: "+time.strftime("%d-%m-%y"))
        
        SONDAGE_PATH= "\SONDAGE"
        if not os.path.exists(SONDAGE_PATH):
                    os.makedirs(SONDAGE_PATH)
        
        
        if not os.path.isfile("RAPPORT_sondage.xlsx)"):
            try:
                s_wb = load_workbook('S_template.xlsx')
                s_wb.save("RAPPORT_sondage.xlsx")
            except Exception as e:
                print("exception permition is denied",e)
                s_wb.close()
                
        s_wb = load_workbook("RAPPORT_sondage.xlsx")    
        s_ws = s_wb.active
        s_ws.print_options.horizontalCentered = True
        #s_ws['G1'] = "Page: 1"
       
        #s_ws['G1'].font = Font(size=18)
        
##        s_ws['A4'] = f"Projet: {PROJECT}"
##        s_ws['A3'] = f'DATE: {time.strftime("%d-%m-%y")}'
##        s_ws['A3'].font = Font(size=18)
        
        s_ws['A3'] ="Fournisseur: "+self.fournisseur.text()
        s_ws['A4'] ="Date:"+ self.date.text()
        s_ws['A5'] ="N° Bobine: "+self.bobine.text()
        
        s_ws['E3'] ="N° de decommande: "+self.commande.text()
        s_ws['E4'] ="Diametre: "+self.diametre.text()
        s_ws['E5'] ="N° de fab: "+self.nemero_fab.text()
        
        s_ws['F3'] ="Installation: "+self.installation.text()
        s_ws['F4'] ="N° coulée: "+self.coulee.text()
        s_ws['F5'] ="Epaisseur : "+self.fournisseur_2.text()
        
        
        s_ws['A27'] ="ANALYSE DES CAUSES:\n"+self.comboBox.currentText()
        s_ws['A30'] ="ACTION CURATIVE IMMEDIATE POUR CORRIGER LE DEFAUT:\n" +self.textEdit.toPlainText()
        s_ws['A33'] ="ACTION CORRECTIVE POUR CORRIGER LA REPARATION DU DEFAUT:\n"+self.textEdit_2.toPlainText()
        
        s_ws['A37'] = "MISE EN OEUVER:                 " + self.MISE_EN_OEUVER
        s_ws['A39'] = "EFFICACITE:                             " + self.EFFICACITE
        s_ws['A41'] = "CLOTURE:                                " + self.CLOTURE
        
        s_ws['A43'] ="DATE ET HEURE:  "+ time.strftime("%d-%m-%y")
        s_ws['E37'] ="NOM ET PRENOM:   "+self.lineEdit_1.text()
        #s_ws['E39'] ="VISA:"+
        s_ws.add_image(xlimg("img2.png"),'F39')
        
        
        
        s_wb.save("RAPPORT_sondage.xlsx")
        print("done!!")
        self.converting_excel_to_pdf(r"{}\RAPPORT_sondage.xlsx".format(P.load_element("path_fab")),
                                     r"{}\RAPPORT_sondage.pdf".format(P.load_element("path_fab")))
        
        self.textEdit.setText("")
        self.textEdit_2.setText("")
        self.checkBox_4.setChecked(False)
        self.checkBox_5.setChecked(False)
        self.checkBox_6.setChecked(False)
        
    

class load_path:
         
    def load_element(self, element):
        self.jsn_list="projet.json"
        try:
          openfile = open(self.jsn_list, 'r')
          jsn_conten = json.load(openfile)
        except:
            outfile = open(self.jsn_list, "w")
            if self.jsn_list=="projet.json":
                J_dict={"projet":"Rgz-2",
                        "Nuance":"x70  MPSL 2",
                        "Diameter":"1016 mm",
                        "Epaisseur":"12,70 mm",
                        "path_fab":r"C:\Users\YASSINE\Desktop\SCRIPTS\fabrication"}
                
                
            json_object = json.dumps(J_dict, indent=2)
            outfile.write(json_object)
            jsn_conten = J_dict
        return jsn_conten[element]
                             
                    
        
            
            
                    
class recievethread(QThread):    
    messge_recieved = pyqtSignal(str)
    messge_to_send= "---+++---"
    def run(self):
        while True:
            try:
                s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                s.connect(('192.168.2.4', 9997))
                
                while True:
                    try:
                        print("got connected")
                        ui.send.connect(self.updatesend)
                        s.send(str(self.messge_to_send).encode('utf-8'))
                        
                        
                        
                        print("FAB", self.messge_to_send)#'test'
                        msg = s.recv(1024)
                        self.messge_recieved.emit(str(msg.decode('utf-8')))
                        
                        ui.listWidget.insertItem(-1,"RX1: "+ str(msg.decode('utf-8')))
                        print("RX1:",msg.decode('utf-8'))
                    except Exception as ex:
                        print("waiting 1 ... ",ex)
                        break
            except Exception as ex:
                 print("waiting 2 ... ",ex)
             
    def updatesend(self,sxt):
        self.messge_to_send = sxt
        #print("updatetd messge_to_send",self.messge_to_send)
       

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    FARICATION = QtWidgets.QMainWindow()
    ui = Ui_FARICATION()
    ui.setupUi(FARICATION)
    P = load_path()
        
    chat  = recievethread()
    chat.start()
    FARICATION.show()
    sys.exit(app.exec_())
    
